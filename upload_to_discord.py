#!/usr/bin/env python3
import os
import glob
import requests
from openpyxl import load_workbook

def upload_report_safe_engine():
    webhook_url = os.environ.get('DISCORD_WEBHOOK_URL')
    if not webhook_url:
        print("[-] Error: DISCORD_WEBHOOK_URL variable is missing.")
        return

    files = glob.glob('reports/passive_recon_report_*.xlsx')
    if not files:
        print("[-] Error: No Excel report asset found in reports/ folder.")
        return
    
    latest_file = max(files, key=os.path.getmtime)
    file_name = os.path.basename(latest_file)
    
    try:
        wb = load_workbook(latest_file, read_only=True)
        total_sheets = len(wb.sheetnames)
        active_domains_count = max(total_sheets - 2, 0)
    except Exception as e:
        print(f"[-] Warning: Failed to parse excel sheet count: {e}")
        active_domains_count = "정상"

    total_new_count = "0"
    if os.path.exists('reports/new_count.txt'):
        try:
            with open('reports/new_count.txt', 'r') as f:
                total_new_count = f.read().strip()
        except: pass

    print(f"[+] Transmitting pure EXCEL workbook ({file_name}) to Discord...", flush=True)
    with open(latest_file, 'rb') as f:
        payload = {
            # 💡 [멘트 변경] 디스코드 봇이 포스트맨 컬렉션 생성 사실을 알려줍니다!
            'content': (
                f"🚀 **[정찰 완료 - 통합 마스터 엑셀 보고서]**\n"
                f"🔥 **새로 발견된 엔드포인트:** `{total_new_count}개`의 신규 URL이 추가되었습니다!\n"
                f"🔒 **{active_domains_count}개 도메인**에 대한 관제 정보가 매핑되었습니다.\n"
                f"📊 아래 엑셀 파일을 터치하여 상세 엔드포인트를 확인하세요.\n"
                f"⚡ **[TIP]** 깃허브 Actions의 다운로드 금고(ZIP) 안에 Burp Suite 연동을 위한 **`Postman Collection (JSON)`** 스펙 파일이 함께 배송되었습니다!"
            )
        }
        files_payload = {'file': (file_name, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(webhook_url, data=payload, files=files_payload)
        
    if response.status_code in [200, 204]:
        print("[+] [SUCCESS] Native single-packet Discord transmission complete!", flush=True)
    else:
        print(f"[-] Discord error code: {response.status_code}, {response.text}", flush=True)

if __name__ == '__main__':
    upload_report_safe_engine()
