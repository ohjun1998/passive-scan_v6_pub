#!/usr/bin/env python3
import os
import glob
import re
import json
import posixpath
from urllib.parse import urlparse, parse_qsl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def escape_formula(value):
    if isinstance(value, str) and value.startswith(('=', '+', '-', '@')): return "'" + value
    return value

def make_absolute(url, domain):
    if url.startswith('http://') or url.startswith('https://'): return url
    elif url.startswith('//'): return f"https:{url}"
    elif url.startswith('/'): return f"https://{domain}{url}"
    else: return f"https://{domain}/{url}"

def get_status_color(status):
    status_str = str(status)
    if status_str.startswith('2'): return '28A745'
    if status_str.startswith('3'): return '17A2B8'
    if status_str.startswith('4'): return 'FD7E14'
    if status_str.startswith('5'): return 'DC3545'
    if 'Static' in status_str: return 'A8B8D0'
    if 'Skipped' in status_str: return 'E83E8C' 
    return '6C757D'

def get_safe_domain(target):
    return "wild_" + target[2:] if target.startswith('*.') else target

def normalize_dynamic_path(path):
    p = re.sub(r'[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}', '{UUID}', path)
    p = re.sub(r'\b\d{3,}\b', '{ID}', p)
    p = re.sub(r'\b[a-zA-Z0-9]{10,}\b', '{HASH}', p)
    return p

def build_advanced_excel_report():
    print("[+] 스마트 네비게이션 및 통계 탑재 대시보드 엔진 가동 중...", flush=True)
    if not os.path.exists('targets.txt'): return
    with open('targets.txt', 'r') as f: targets = [line.strip() for line in f if line.strip()]

    target_map = {get_safe_domain(t): t for t in targets}
    js_url_converter = {}
    for mf in glob.glob('results/*_js_mapping.txt'):
        try:
            with open(mf, 'r', errors='ignore') as f:
                for line in f:
                    if '\t' in line:
                        s, o = line.strip().split('\t', 1)
                        js_url_converter[s] = o
        except: pass

    matrix_data = {raw_target: {} for raw_target in targets}
    signature_counts = {}
    
    junk_extensions = ('.png', '.jpg', '.jpeg', '.gif', '.svg', '.css', '.woff', '.woff2', '.ico', '.eot', '.ttf', '.mp4')
    blacklist_words = ['logout', 'signout', 'delete', 'remove', 'revoke', 'destroy']
    
    for file_path in glob.glob('results/*.*'):
        filename = os.path.basename(file_path).lower()
        match = re.match(r'^(.*)_(linkfinder|trufflehog|gau|waybackurls)\.txt$', filename)
        if not match: continue
        
        safe_domain = match.group(1)
        if safe_domain not in target_map: continue
        
        raw_target = target_map[safe_domain]
        is_wildcard = raw_target.startswith('*.')
        base_domain = raw_target[2:] if is_wildcard else raw_target

        if 'linkfinder' in filename or 'jsluice' in filename: source_tool = 'LinkFinder'
        elif 'trufflehog' in filename: source_tool = 'TruffleHog'
        elif 'waybackurls' in filename: source_tool = 'Waybackurls'
        elif 'gau' in filename: source_tool = 'GAU'
        else: continue

        try:
            with open(file_path, 'r', errors='ignore') as f:
                for line in f:
                    line_str = line.strip()
                    if not line_str or line_str.startswith('#'): continue
                    line_str = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', line_str)
                    if not line_str: continue
                    
                    if '\t' in line_str: js_file, raw_url = line_str.split('\t', 1)
                    else: js_file, raw_url = "Passive Archive", line_str
                    
                    if js_file in js_url_converter: js_file = js_url_converter[js_file]
                    
                    abs_url = make_absolute(raw_url, base_domain)
                    parsed_netloc = urlparse(abs_url).netloc.split(':')[0]
                    
                    if is_wildcard:
                        if not (parsed_netloc == base_domain or parsed_netloc.endswith('.' + base_domain)):
                            continue
                    else:
                        if parsed_netloc != base_domain:
                            continue

                    if urlparse(abs_url).path.lower().endswith(junk_extensions): continue

                    parsed_for_sig = urlparse(abs_url)
                    query_keys = tuple(sorted([k for k, v in parse_qsl(parsed_for_sig.query, keep_blank_values=True)]))
                    norm_path = normalize_dynamic_path(parsed_for_sig.path)
                    path_dir = posixpath.dirname(norm_path)
                    path_ext = posixpath.splitext(norm_path)[1]
                    signature = (parsed_for_sig.netloc, path_dir, path_ext, query_keys)

                    if abs_url not in matrix_data[raw_target]:
                        if signature_counts.get(signature, 0) >= 5:
                            continue 
                        signature_counts[signature] = signature_counts.get(signature, 0) + 1
                        matrix_data[raw_target][abs_url] = {"tools": set(), "files": set()}
                        
                    matrix_data[raw_target][abs_url]["tools"].add(source_tool)
                    if source_tool in ['LinkFinder', 'TruffleHog']:
                        matrix_data[raw_target][abs_url]["files"].add(js_file)
        except: pass

    status_codes = {}
    for res_file in glob.glob('results/httpx_results_*.json'):
        try:
            with open(res_file, 'r', errors='ignore') as f:
                for line in f:
                    data = json.loads(line.strip())
                    status_codes[data.get('url')] = data.get('status_code', 'Dead')
        except: pass

    wb = Workbook()
    font_header, fill_header = Font(name='Malgun Gothic', bold=True, color='FFFFFF'), PatternFill(start_color='2F3542', end_color='2F3542', fill_type='solid')
    font_data, fill_zebra = Font(name='Malgun Gothic', size=10, color='333333'), PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    align_center, align_left = Alignment(horizontal='center', vertical='center'), Alignment(horizontal='left', vertical='center')
    thin_border = Border(left=Side(style="thin", color="E0E0E0"), right=Side(style="thin", color="E0E0E0"), top=Side(style="thin", color="E0E0E0"), bottom=Side(style="thin", color="E0E0E0"))

    # ==========================================
    # 1. Summary Dashboard 생성
    # ==========================================
    ws_dash = wb.active
    ws_dash.title = "Summary Dashboard"
    ws_dash.append(["No", "타겟 도메인", "완전 정제된 URL 수", "jsluice 추출 개수", "TruffleHog 탐지 개수"])
    for c in range(1, 6): ws_dash.cell(1, c).font = font_header; ws_dash.cell(1, c).fill = fill_header; ws_dash.cell(1, c).alignment = align_center; ws_dash.cell(1, c).border = thin_border

    # ==========================================
    # 2. High Risk Targets 시트 세팅 (네비게이션 버튼 포함)
    # ==========================================
    ws_high = wb.create_sheet(title="High Risk Targets")
    
    ws_high.append(["🔙 대시보드로 돌아가기 (Return to Dashboard)"])
    ws_high.merge_cells('A1:G1')
    back_cell_h = ws_high.cell(row=1, column=1)
    back_cell_h.hyperlink = "#'Summary Dashboard'!A1"
    back_cell_h.font = Font(name='Malgun Gothic', size=11, bold=True, color='0056B3', underline='single')
    back_cell_h.fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
    back_cell_h.alignment = align_left

    ws_high.append(["No", "소스 출처", "발견된 JS 파일명", "응답 상태", "도메인", "고위험 경로 (Endpoint)", "탐지 사유"]) 
    for c in range(1, 8): ws_high.cell(2, c).font = font_header; ws_high.cell(2, c).fill = fill_header; ws_high.cell(2, c).alignment = align_center; ws_high.cell(2, c).border = thin_border

    # 💡 [버그 수정 완료] 에러의 원인이었던 누락된 고위험 키워드 리스트 복원
    high_risk_keywords = ['config', '.env', 'xml', 'json', 'secret', 'api/v', 'token', 'admin', 'password', 'key', 'credential', 'mysql']
    
    dash_idx = 2
    high_risk_idx = 3

    # ==========================================
    # 3. 개별 도메인별 데이터 삽입 및 시트 생성
    # ==========================================
    for raw_target, url_map in matrix_data.items():
        sheet_title = re.sub(r'[\\/\?\*\:\[\]]', '_', raw_target)[:30]
        passive_count = sum(1 for data in url_map.values() if 'Waybackurls' in data["tools"] or 'GAU' in data["tools"])
        jsluice_count = sum(1 for data in url_map.values() if 'LinkFinder' in data["tools"])
        trufflehog_count = sum(1 for data in url_map.values() if 'TruffleHog' in data["tools"])
        
        ws_dash.append([dash_idx - 1, escape_formula(raw_target), passive_count, jsluice_count, trufflehog_count])
        for c in range(1, 6):
            ws_dash.cell(dash_idx, c).font = font_data; ws_dash.cell(dash_idx, c).border = thin_border
            ws_dash.cell(dash_idx, c).alignment = align_left if c == 2 else align_center
            if c == 2 and url_map: ws_dash.cell(dash_idx, c).hyperlink = f"#'{sheet_title}'!A1"; ws_dash.cell(dash_idx, c).font = Font(name='Malgun Gothic', color='0056B3', underline='single')
        dash_idx += 1

        if not url_map: continue

        ws = wb.create_sheet(title=sheet_title)
        
        ws.append(["🔙 대시보드로 돌아가기 (Return to Dashboard)"])
        ws.merge_cells('A1:E1')
        back_cell = ws.cell(row=1, column=1)
        back_cell.hyperlink = "#'Summary Dashboard'!A1"
        back_cell.font = Font(name='Malgun Gothic', size=11, bold=True, color='0056B3', underline='single')
        back_cell.fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
        back_cell.alignment = align_left

        ws.append(["No", "소스 출처", "발견된 JS 파일명", "응답 상태", "타겟 절대 경로 (URL)"]) 
        for c in range(1, 6): ws.cell(2, c).font = font_header; ws.cell(2, c).fill = fill_header; ws.cell(2, c).alignment = align_center; ws.cell(2, c).border = thin_border

        for sub_idx, (url, data) in enumerate(sorted(url_map.items()), 1):
            if sub_idx > 1048500: break
            tools_str = ", ".join(sorted(list(data["tools"])))
            files_str = ", ".join(sorted(list(data["files"]))) if data["files"] else "-"
            
            is_blacklist = any(b in url.lower() for b in blacklist_words)
            
            if is_blacklist:
                current_status = "Skipped(위험)"
            elif urlparse(url).path.lower().endswith(junk_extensions):
                current_status = "Static(생략)"
            else:
                current_status = status_codes.get(url, 'Dead')
            
            row_num = sub_idx + 2
            ws.append([sub_idx, escape_formula(tools_str), escape_formula(files_str), current_status, escape_formula(url)]) 
            
            for c in range(1, 6):
                cell = ws.cell(row_num, c)
                cell.font = font_data; cell.border = thin_border
                if (row_num % 2) == 1: cell.fill = fill_zebra
                
                if c == 4:
                    cell.fill = PatternFill(start_color=get_status_color(current_status), end_color=get_status_color(current_status), fill_type='solid')
                    cell.font = Font(name='Malgun Gothic', bold=True, color='FFFFFF'); cell.alignment = align_center
                elif c in [3, 5]: cell.alignment = align_left
                else: cell.alignment = align_center

            is_high_risk, reason = False, ""
            if 'TruffleHog' in data["tools"]: 
                is_high_risk, reason = True, "TruffleHog 검증 완료: 민감 키(Secret) 유출 징후 탐지"
            elif is_blacklist:
                is_high_risk, reason = True, "파괴적 엔드포인트 (Httpx 스캔 제외 및 수동 점검 요망)"
            else:
                matched = [key for key in high_risk_keywords if key in url.lower()]
                if matched: is_high_risk, reason = True, f"민감 키워드 감지 ({', '.join(matched)})"
                    
            if is_high_risk:
                ws_high.append([high_risk_idx - 2, escape_formula(tools_str), escape_formula(files_str), current_status, escape_formula(raw_target), escape_formula(url), escape_formula(reason)]) 
                for c in range(1, 8):
                    cell = ws_high.cell(high_risk_idx, c)
                    cell.font = font_data; cell.border = thin_border
                    if (high_risk_idx % 2) == 1: cell.fill = fill_zebra
                    if c == 4:
                        cell.fill = PatternFill(start_color=get_status_color(current_status), end_color=get_status_color(current_status), fill_type='solid')
                        cell.font = Font(name='Malgun Gothic', bold=True, color='FFFFFF'); cell.alignment = align_center
                    elif c in [3, 5, 6, 7]: cell.alignment = align_left
                    else: cell.alignment = align_center
                high_risk_idx += 1

    # ==========================================
    # 💡 대시보드 하단 총 합계(Total) 계산 수식 삽입
    # ==========================================
    if dash_idx > 2:
        ws_dash.append(["", "📊 총 합계 (Total)", f"=SUM(C2:C{dash_idx-1})", f"=SUM(D2:D{dash_idx-1})", f"=SUM(E2:E{dash_idx-1})"])
        for c in range(1, 6):
            cell = ws_dash.cell(dash_idx, c)
            cell.font = Font(name='Malgun Gothic', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            cell.border = thin_border
            cell.alignment = align_center if c != 2 else align_left

    # 컬럼 자동 너비 조절 및 최적화
    for sheet in wb.worksheets:
        header_row = 1 if sheet.title == "Summary Dashboard" else 2
        for col_idx, col in enumerate(sheet.columns, 1):
            col_letter = get_column_letter(col_idx)
            header_cell_value = sheet.cell(header_row, col_idx).value
            header = str(header_cell_value) if header_cell_value else ""
            
            if header in ["타겟 절대 경로 (URL)", "고위험 경로 (Endpoint)"]: sheet.column_dimensions[col_letter].width = 80  
            elif header == "발견된 JS 파일명": sheet.column_dimensions[col_letter].width = 50  
            elif header == "탐지 사유": sheet.column_dimensions[col_letter].width = 40  
            elif header == "응답 상태": sheet.column_dimensions[col_letter].width = 16
            else: sheet.column_dimensions[col_letter].width = 18

    ws_dash.column_dimensions['B'].width = 35
    wb.active = 0 
    
    os.makedirs('reports', exist_ok=True)
    wb.save('reports/passive_recon_report_v1.xlsx')
    print("[+] 네비게이션 적용 엑셀 보고서 렌더링이 성공적으로 완료되었습니다!", flush=True)

if __name__ == '__main__':
    build_advanced_excel_report()
